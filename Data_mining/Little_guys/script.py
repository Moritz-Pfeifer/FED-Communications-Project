from selenium import webdriver
from selenium_stealth import stealth
import sys
import os
from time import sleep
import requests
from selenium.webdriver.support.ui import WebDriverWait as WDW
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from openpyxl import Workbook
from bs4 import BeautifulSoup
from openpyxl import load_workbook


wb = Workbook()
ws = wb.active

options = webdriver.ChromeOptions()
options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36")
options.add_experimental_option("excludeSwitches", ["enable-automation"])
options.add_experimental_option('useAutomationExtension', False)
options.add_argument("--log-level=3")
options.add_argument('--disable-blink-features=AutomationControlled')
driver = webdriver.Chrome(options=options, executable_path=os.path.join(sys.path[0], "chromedriver"))
stealth(driver,
        languages=["en-US", "en"],
        vendor="Google Inc.",
        platform="Win32",
        webgl_vendor="Intel Inc.",
        renderer="Intel Iris OpenGL Engine",
        fix_hairline=True,)

def waitfor(xpath,err):
    try:
        WDW(driver, 20).until(EC.presence_of_element_located((By.XPATH, xpath)))
    except:
        print(err)
        pass


def window_handles(window_number: int) -> None:

    driver.switch_to.window(driver.window_handles[window_number])

def scroll(element):
    desired_y = (element.size['height'] / 2) + element.location['y']
    window_h = driver.execute_script('return window.innerHeight')
    window_y = driver.execute_script('return window.pageYOffset')
    current_y = (window_h / 2) + window_y
    scroll_y_by = desired_y - current_y
    driver.execute_script("window.scrollBy(0, arguments[0]);", scroll_y_by)

print("ok")

def richmondfed():
    row = 1
    url = "https://www.richmondfed.org/press_room/speeches/j_alfred_broaddus/archive/#2 "
    driver.get(url)
    texts = driver.find_elements_by_xpath('/html/body/div[1]/div[2]/div/div/div[2]/div[2]/div[1]/ul/li/div/div')
    for text in texts:
        date = text.find_element_by_xpath('./section/div[1]').text
        titlee = text.find_element_by_xpath('./section/div[2]')
        title = titlee.text
        link = titlee.find_element_by_xpath('./a').get_attribute('href')
        summary = text.find_element_by_xpath('./section/div[3]').text
        try:
            author = text.find_element_by_xpath('./section/div[4]').text
        except:
            author = summary
            summary = ""
        if "barkin" in author.lower():
            ws['A' + str(row)].value = author
            ws['B' + str(row)].value = date
            ws['C' + str(row)].value = title
            ws['D' + str(row)].value = link
            ws['E' + str(row)].value = summary
            row += 1

    wb.save("TomBarkin.xlsx")


def chicagofed():
    row = 1
    url = "https://www.chicagofed.org/publications/speeches/president-events"
    driver.get(url)
    for i in range(2015,2023):
        years = driver.find_elements_by_id('cfedJL-' + str(i))
        for year in years:
            lines = year.find_elements_by_xpath('./div')
            for line in lines:
                try:
                    titlee = line.find_element_by_xpath('./div[1]/a')
                    title = titlee.text
                    link = titlee.get_attribute('href')
                except:
                    title = line.find_element_by_xpath('./div[1]').text
                    link = "None"
                summary = line.find_element_by_xpath('./div[2]').text
                date = line.find_element_by_xpath('./div[3]').text + " " + str(i)
                ws['A' + str(row)].value = "Charles.L.Evans"
                ws['B' + str(row)].value = date
                ws['C' + str(row)].value = title
                ws['D' + str(row)].value = link
                ws['E' + str(row)].value = summary
                row += 1

    wb.save("CharlesLEvansNew.xlsx")


def dallasfed():
    row = 1
    url = 'https://www.dallasfed.org/news/speeches/fisher.aspx'
    driver.get(url)
    blocks = driver.find_elements_by_xpath('/html/body/div[3]/div[1]/div[2]/div/div[1]/div[3]/div/p')
    for block in blocks:
        list = block.text.split("\n")
        if len(list) > 1:
            print(list)
            ws['A' + str(row)].value = "Richard W. Fisher"
            try:
                ws['B' + str(row)].value = block.find_element_by_xpath('./strong/a').get_attribute('href')
            except:
                try:
                    ws['B' + str(row)].value = block.find_element_by_xpath('./a').get_attribute('href')
                except:
                    ws['B' + str(row)].value = block.find_element_by_xpath('./b/a').get_attribute('href')
            for item in list:
                ws.cell(row=row ,column=list.index(item) + 3).value = item
            row += 1

    wb.save("RichardWFisher.xlsx")
    input('ok')


def frbsf():
    row = 1
    url = 'https://www.frbsf.org/our-district/press/presidents-speeches/parry-speeches/#2003'
    driver.get(url)
    blocks = driver.find_elements_by_xpath('/html/body/div[2]/div[2]/div[2]/div/main/article/section/article/div')
    for block in blocks:
        titlee = block.find_element_by_xpath('./h2/a')
        title = titlee.text
        link = titlee.get_attribute('href')
        summary = block.find_element_by_xpath('./p[2]').text
        date = block.find_element_by_xpath('./p[3]').text
        ws['A' + str(row)].value = "Mary C. Daly"
        ws['B' + str(row)].value = date
        ws['C' + str(row)].value = title
        ws['D' + str(row)].value = link
        ws['E' + str(row)].value = summary
        row +=1
    wb.save('MaryCDaly.xlsx')


def atlantafed():
    row = 1
    url = "https://www.atlantafed.org/news/speeches/?pub_year="
    for i in range(1995,2022):
        driver.get(url + str(i))
        ps = driver.find_elements_by_xpath('/html/body/div[1]/div/div/div/div[2]/div/div/div/div[2]/div[1]/div/span/p')
        for o in range(0,len(ps)-1):
            if o%3 == 0:
                date = ps[o].text
                title = ps[o+1].text
                link = ps[o+1].find_element_by_xpath('./a').get_attribute('href')
                summary = ps[o+2].text
                if "bostic" in summary.lower():
                    ws['A' + str(row)].value = title
                    ws['B' + str(row)].value = date
                    ws['C' + str(row)].value = link
                    ws['D' + str(row)].value = summary
                    row += 1
        wb.save("Bostic.xlsx")
        break

def kansascityfed():
    row = 1
    url = "https://www.kansascityfed.org/search/?filter-section=page&sorting=relevant&page-filter-category=6&6-12=6004&page-number=1&perpage=100"
    driver.get(url)
    blocks = driver.find_elements_by_xpath('/html/body/main/div[1]/section[2]/div/div[2]/div/div/div/div[2]')
    for block in blocks:
        titlee = block.find_element_by_xpath('./h3/a')
        title = titlee.text
        link = titlee.get_attribute('href')
        date = block.find_element_by_xpath('./span').text
        author = block.find_element_by_xpath('./div/ul/li[1]/ul/li/a').text
        ws['A' + str(row)].value = author
        ws['B' + str(row)].value = title
        ws['C' + str(row)].value = date
        ws['D' + str(row)].value = link
        row += 1

    wb.save('EstherLGeorge.xlsx')


def minneapolisfed():
    row = 1
    url = 'https://www.minneapolisfed.org/people/narayana-kocherlakota'
    driver.get(url)
    blocks = driver.find_elements_by_xpath('/html/body/main/div/div[2]/div/div/div[2]/div[2]/ul/li')
    blocks = blocks + driver.find_elements_by_xpath('/html/body/main/div/div[2]/div/div/div[2]/div[2]/ul/div/li')
    for block in blocks:
        titlee = block.find_element_by_xpath('./a')
        title = titlee.text
        print(title)
        link = titlee.get_attribute('href')
        date =block.find_element_by_xpath('./div/div[1]').text
        type = block.find_element_by_xpath('./div/div[2]').text
        if "speech" in type.lower() and 'kashkari' in title.lower():
            ws['A' + str(row)].value = "Neel Kashkari"
            ws['B' + str(row)].value = title
            ws['C' + str(row)].value = date
            ws['D' + str(row)].value = link
            row += 1

    wb.save('Neel Kashkari.xlsx')

def clevelandfed():
    row = 1
    wb = Workbook()
    ws = wb.active

    url = 'https://www.clevelandfed.org/newsroom-and-events/speeches.aspx'
    driver.get(url)
    blocks = driver.find_elements_by_xpath('/html/body/div[2]/div[2]/div[2]/div[3]/div/ul/li')
    for block in blocks:
        author = block.find_element_by_xpath('./p[1]').text
        if 'mester' in author.lower():
            titlee = block.find_element_by_xpath('./h4/a')
            title = titlee.text
            link = titlee.get_attribute('href')
            date = block.find_element_by_xpath('./time').text
            summary = block.find_element_by_xpath('./p[3]').text

            ws['A' + str(row)].value = "Loretta Mester"
            ws['B' + str(row)].value = title
            ws['C' + str(row)].value = date
            ws['D' + str(row)].value = link
            ws['E' + str(row)].value = summary
            row += 1
    wb.save('LorettaMester.xlsx')

def newyorksfed():
    row = 1
    wb = Workbook()
    ws = wb.active

    url = 'https://www.newyorkfed.org/aboutthefed/orgchart/williams'
    driver.get(url)

    blocks = driver.find_elements_by_xpath('/html/body/div[2]/div[6]/div[7]/div')
    for block in blocks:
        if blocks.index(block) is not 0:
            titlee = block.find_element_by_xpath('./div[1]/a')
            title = titlee.text
            link = titlee.get_attribute('href')
            date = block.find_element_by_xpath('./div[2]').text

            ws['A' + str(row)].value = "John C. Williams"
            ws['B' + str(row)].value = title
            ws['C' + str(row)].value = date
            ws['D' + str(row)].value = link
            row += 1

    wb.save('JohnCWilliams.xlsx')


def fraser():
    row = 1
    wb = Workbook()
    ws = wb.active
    url = 'https://fraser.stlouisfed.org/author/baughman-ernest-t '
    driver.get(url)

    blocks = driver.find_elements_by_xpath('/html/body/div/div[3]/div[6]/div[1]/div/div[1]/div[3]/ul[1]/li')
    for block in blocks:
        scroll(block)
        block.click()
        waitfor('/html/body/div/div[3]/div[6]/div[1]/div/div[2]/h2/a', 'no box')
        title = driver.find_element_by_xpath('/html/body/div/div[3]/div[6]/div[1]/div/div[2]/h2/a').text
        print(title)
        link = driver.find_element_by_xpath('/html/body/div/div[3]/div[6]/div[1]/div/div[2]/h2/a').get_attribute('href')
        date = driver.find_element_by_xpath('/html/body/div/div[3]/div[6]/div[1]/div/div[2]/p[1]/span[2]').text
        author = "Ernest T. Baughman"

        ws['A' + str(row)].value = author
        ws['B' + str(row)].value = title
        ws['C' + str(row)].value = date
        ws['D' + str(row)].value = link
        row += 1

    wb.save(author + '.xlsx')


