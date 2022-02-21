from concurrent.futures import as_completed
from requests_futures.sessions import FuturesSession
from concurrent.futures import ThreadPoolExecutor
import json
from bs4 import BeautifulSoup
import requests
from openpyxl import load_workbook
import re

url = "https://fraser.stlouisfed.org"
wb = load_workbook('./test2.xlsx')
ws = wb.active
done = []
last = ws.max_row

for row in range(1, last+1):
    try:
        link = ws["F" + str(row)].value
        if title not in done:
            done.append(title)
        else:
            ws.delete_rows(row)
    except:
        print(row)
        pass

wb.save("delete.xlsx")
