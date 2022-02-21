from openpyxl import load_workbook
import calendar

wb = load_workbook('afterleng.xlsx')
ws = wb.active
last = ws.max_row

def lengg():
    for row in range(1, last+1):
        text = ws.cell(row=row, column=6).value
        if text is not None:
            leng = len(text.split(" "))
            print(row)
            ws.cell(row=row, column=8).value = leng

    wb.save('afterleng.xlsx')

def year():
    for row in range(1, last+1):
        date = ws.cell(row=row, column=5).value
        if date is not None:
            year = date.split(" ")[-1]
            print(year)
            if '.' in year:
                year = '20' + year.split('.')[0]
            ws.cell(row=row, column=5).value = int(year)

    wb.save('fraserbruuuh.xlsx')

def date():
    for row in range(1, last+1):
        try:
            date = ws.cell(row=row, column=7).value
            if date is not None:
                date = date.replace("Speech on ", '')
                split = date.split(" ")
                if len(split) == 3:
                    year = split[2]
                    month = list(calendar.month_name).index(split[0].replace('.',''))
                    day = split[1].replace(",", "")
                    new = f'{year}.{month}.{day}'
                    ws.cell(row=row, column=7).value = new
        except:
            print(date)

    wb.save('dooooooednoadeaod.xlsx')

states = ['Alabama', 'Alaska', 'Arizona', 'Arkansas', 'California', 'Colorado', 'Connecticut', 'Delaware', 'Florida', 'Georgia', 'Hawaii', 'Idaho', 'Illinois', 'Indiana', 'Iowa', 'Kansas', 'Kentucky', 'Louisiana', 'Maine', 'Maryland', 'Massachusetts', 'Michigan', 'Minnesota', 'Mississippi', 'Missouri', 'Montana', 'Nebraska', 'Nevada', 'New Hampshire', 'New Jersey', 'New Mexico', 'New York', 'North Carolina', 'North Dakota', 'Ohio', 'Oklahoma', 'Oregon', 'Pennsylvania', 'Rhode Island', 'South Carolina', 'South Dakota', 'Tennessee', 'Texas', 'Utah', 'Vermont', 'Virginia', 'Washington', 'West Virginia', 'Wisconsin', 'Wyoming', 'Washington D.C']

def location():
    for row in range(1, last+1):
        text = ws.cell(row=row, column=6).value
        if text is not None:
            found = []
            for state in states:
                if state in text:
                    found.append(state)

            sort_found = sorted(found, key=text.find)
            if len(sort_found) != 0:
                ws.cell(row=row, column=9).value = sort_found[0]
            else:
                print(row)

    wb.save('afterloc.xlsx')

location()